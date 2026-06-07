"""
Report service - generates CSV and Excel reports using Pandas and OpenPyXL.
"""
import pandas as pd
import re
from io import BytesIO
from app.models.task import Task
from app.models.attendance import Attendance
from app.models.user import User, UserRole
from beanie import PydanticObjectId
from datetime import datetime, timezone, timedelta, time
from typing import Optional, List

def parse_time_string(time_str: Optional[str]) -> time:
    """
    Parse check-in / start time strings into a datetime.time object.
    Supports formats:
      - 12-hour: '09:00 AM', '9:30PM', '09:00AM', etc.
      - 24-hour: '14:30', '09:00', '9:00', etc.
    Falls back to time(9, 0) on any exception or mismatch.
    """
    if not time_str:
        return time(9, 0)
    
    raw = str(time_str).strip().upper()
    
    # Try Regex for 12-hour format: e.g. 09:30 AM or 9:30PM
    m12 = re.match(r'^(\d{1,2}):(\d{2})\s*(AM|PM)$', raw)
    if m12:
        h, m, p = m12.groups()
        h, m = int(h), int(m)
        if p == "PM" and h < 12:
            h += 12
        elif p == "AM" and h == 12:
            h = 0
        try:
            return time(h, m)
        except ValueError:
            return time(9, 0)
            
    # Try Regex for 24-hour format: e.g. 14:30 or 09:00
    m24 = re.match(r'^(\d{1,2}):(\d{2})$', raw)
    if m24:
        h, m = m24.groups()
        h, m = int(h), int(m)
        try:
            return time(h, m)
        except ValueError:
            return time(9, 0)

    # Strptime fallbacks
    formats = ["%I:%M %p", "%I:%M%p", "%H:%M", "%I:%M  %p"]
    for fmt in formats:
        try:
            dt = datetime.strptime(raw, fmt)
            return dt.time()
        except ValueError:
            continue
            
    return time(9, 0)
try:
    from zoneinfo import ZoneInfo
except ImportError:
    from backports.zoneinfo import ZoneInfo


async def _get_task_data(
    current_user: User,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    priority: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    tz_name: Optional[str] = None,
) -> pd.DataFrame:
    """Fetch and filter task data into a DataFrame with hierarchy-based filtering."""
    from app.services import user_service

    query = {}

    from app.models.role import BaseArchetype
    arch = current_user.role_archetype or current_user.role

    if arch in [BaseArchetype.SUPER_ADMIN, UserRole.SUPER_ADMIN]:
        if employee_id:
            query["assigned_to"] = PydanticObjectId(employee_id)
    elif arch in [BaseArchetype.ADMIN, UserRole.ADMIN, BaseArchetype.HR, BaseArchetype.FINANCE, BaseArchetype.AUDITOR]:
        from app.models.company import Company
        companies = await Company.find({"$or": [{"owner_id": current_user.id}, {"_id": current_user.company_id}]}).to_list()
        co_ids = [c.id for c in companies]
        query["company_id"] = {"$in": co_ids}
        if employee_id:
            query["assigned_to"] = PydanticObjectId(employee_id)
    elif arch in [BaseArchetype.MANAGER, BaseArchetype.ASSISTANT_MANAGER, UserRole.MANAGER, UserRole.ASSISTANT_MANAGER]:
        subordinates = await user_service.get_all_employees(current_user)
        subordinate_ids = {emp.id for emp in subordinates}
        if employee_id:
            target_uid = PydanticObjectId(employee_id)
            if target_uid != current_user.id and target_uid not in subordinate_ids:
                return pd.DataFrame()
            query["assigned_to"] = target_uid
        else:
            query["$or"] = [
                {"assigned_to": {"$in": list(subordinate_ids) + [current_user.id]}},
                {"created_by": current_user.id}
            ]
    else:  # EMPLOYEE, CONTRACTOR
        query["assigned_to"] = current_user.id

    if status:
        query["status"] = status
    if priority:
        query["priority"] = priority
    if start_date:
        query["created_at"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "created_at" in query:
            query["created_at"]["$lte"] = datetime.fromisoformat(end_date)
        else:
            query["created_at"] = {"$lte": datetime.fromisoformat(end_date)}

    tasks = await Task.find(query).sort("-created_at").to_list()

    rows = []
    # Determine timezone for formatting
    tz = ZoneInfo(tz_name) if tz_name else None

    def fmt_dt(dt: datetime) -> str:
        """Format datetime in local timezone if provided."""
        if dt is None:
            return ""
        if tz:
            dt = dt.replace(tzinfo=timezone.utc).astimezone(tz)
        return dt.strftime("%d-%m-%Y %H:%M:%S")

    for i, task in enumerate(tasks, 1):
        # Calculate Time Variance (Deadline - Completed Time)
        time_variance = ""
        if task.completed_at:
            variance = task.deadline - task.completed_at
            hours = variance.total_seconds() / 3600
            if hours > 0:
                time_variance = f"{hours:.1f}h Early"
            else:
                time_variance = f"{abs(hours):.1f}h Late"

        # Format Remarks (Join all remark texts)
        remarks_str = " | ".join([r.get("text", "") for r in task.remarks]) if task.remarks else ""

        rows.append({
            "s.no": i,
            "employee name": task.assigned_to_name or "Unknown",
            "company name": task.company_name or "Personal / Internal",
            "category": ", ".join(task.category_names) if task.category_names else "",
            "work description": task.work_description,
            "work priority": task.priority.value.capitalize(),
            "dead-line": fmt_dt(task.deadline),
            "completed time": fmt_dt(task.completed_at) if task.completed_at else "",
            "Time variance": time_variance,
            "Status": task.status.value.capitalize(),
            "Remarks": remarks_str,
            "points": 1 if task.status == "completed" else 0,
            "created time": fmt_dt(task.created_at),
            "Assigned by": task.created_by_name or "Unknown"
        })

    df = pd.DataFrame(rows)
    # Convert all column names to UPPERCASE as per requirement
    df.columns = [str(c).upper() for c in df.columns]
    return df


async def generate_tasks_csv(
    current_user: User,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    priority: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    tz_name: Optional[str] = None,
) -> str:
    """Generate CSV string of task data."""
    df = await _get_task_data(current_user, status, employee_id, priority, start_date, end_date, tz_name)
    return df.to_csv(index=False)


async def generate_tasks_excel(
    current_user: User,
    status: Optional[str] = None,
    employee_id: Optional[str] = None,
    priority: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    tz_name: Optional[str] = None,
) -> BytesIO:
    """Generate Excel file of task data."""
    df = await _get_task_data(current_user, status, employee_id, priority, start_date, end_date, tz_name)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Tasks", index=False)
    output.seek(0)
    return output


async def generate_employees_excel(current_user: User) -> BytesIO:
    """Generate Excel file of employee data with reward info (filtered by hierarchy)."""
    from app.services import user_service
    employees = await user_service.get_all_employees(current_user)
    # Sort by reward points descending
    employees.sort(key=lambda x: x.reward_points or 0, reverse=True)

    emp_ids = [emp.id for emp in employees]

    # Batch fetch task stats using aggregation
    pipeline = [
        {"$match": {"assigned_to": {"$in": emp_ids}}},
        {"$group": {
            "_id": "$assigned_to",
            "total_tasks": {"$sum": 1},
            "completed_tasks": {
                "$sum": {"$cond": [{"$eq": ["$status", "completed"]}, 1, 0]}
            }
        }}
    ]
    cursor = Task.aggregate(pipeline)
    results = await cursor.to_list(length=None)

    task_stats = {
        res["_id"]: {"total": res["total_tasks"], "completed": res["completed_tasks"]}
        for res in results
    }

    rows = []
    for emp in employees:
        stats = task_stats.get(emp.id, {"total": 0, "completed": 0})
        total_tasks = stats["total"]
        completed_tasks = stats["completed"]

        rows.append({
            "Employee ID": str(emp.id),
            "Name": emp.name,
            "Email": emp.email,
            "Status": "Active" if emp.is_active else "Inactive",
            "Reward Points": emp.reward_points,
            "Total Tasks": total_tasks,
            "Completed Tasks": completed_tasks,
            "Completion Rate": f"{(completed_tasks / total_tasks * 100):.1f}%" if total_tasks > 0 else "0%",
            "Joined": emp.created_at.strftime("%d-%m-%Y %H:%M:%S") if emp.created_at else "",
        })

    df = pd.DataFrame(rows)
    # Convert all column names to UPPERCASE as per requirement
    df.columns = [str(c).upper() for c in df.columns]
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Employees", index=False)
    output.seek(0)
    return output


async def generate_attendance_excel(
    current_user: User,
    user_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    tz_name: Optional[str] = None,
) -> BytesIO:
    """Generate Excel file of attendance data (filtered by hierarchy)."""
    from app.services import user_service
    query = {}

    from app.models.role import BaseArchetype
    arch = current_user.role_archetype or current_user.role

    if arch in [BaseArchetype.SUPER_ADMIN, UserRole.SUPER_ADMIN]:
        if user_id:
            query["user_id"] = PydanticObjectId(user_id)
    elif arch in [BaseArchetype.ADMIN, UserRole.ADMIN, BaseArchetype.HR, BaseArchetype.FINANCE, BaseArchetype.AUDITOR]:
        from app.models.company import Company
        companies = await Company.find({"$or": [{"owner_id": current_user.id}, {"_id": current_user.company_id}]}).to_list()
        co_ids = [c.id for c in companies]
        users = await User.find({"company_id": {"$in": co_ids}}).to_list()
        user_ids = [u.id for u in users] + [current_user.id]
        if user_id:
            target_uid = PydanticObjectId(user_id)
            if target_uid not in user_ids:
                return BytesIO()
            query["user_id"] = target_uid
        else:
            query["user_id"] = {"$in": user_ids}
    elif arch in [BaseArchetype.MANAGER, BaseArchetype.ASSISTANT_MANAGER, UserRole.MANAGER, UserRole.ASSISTANT_MANAGER]:
        subordinates = await user_service.get_all_employees(current_user)
        subordinate_ids = {emp.id for emp in subordinates}
        if user_id:
            target_uid = PydanticObjectId(user_id)
            if target_uid != current_user.id and target_uid not in subordinate_ids:
                return BytesIO()
            query["user_id"] = target_uid
        else:
            query["user_id"] = {"$in": list(subordinate_ids) + [current_user.id]}
    else:
        query["user_id"] = current_user.id
    
    if start_date:
        query["check_in"] = {"$gte": datetime.fromisoformat(start_date)}
    if end_date:
        if "check_in" in query:
            query["check_in"]["$lte"] = datetime.fromisoformat(end_date)
        else:
            query["check_in"] = {"$lte": datetime.fromisoformat(end_date)}

    records = await Attendance.find(query).sort("-check_in").to_list()

    # Pre-fetch user names if needed
    user_map = {}
    if not user_id:
        user_ids = list(set([r.user_id for r in records]))
        users = await User.find({"_id": {"$in": user_ids}}).to_list()
        user_map = {u.id: {"name": u.name, "email": u.email} for u in users}

    # Determine timezone for formatting
    tz = ZoneInfo(tz_name) if tz_name else None

    def to_local(dt: datetime) -> datetime:
        """Convert a UTC datetime to local timezone."""
        if dt is None:
            return None
        if tz:
            return dt.replace(tzinfo=timezone.utc).astimezone(tz)
        return dt

    rows = []
    for i, rec in enumerate(records, 1):
        local_check_in = to_local(rec.check_in)
        local_check_out = to_local(rec.check_out) if rec.check_out else None

        # Calculate duration if checked out
        duration_str = ""
        if rec.check_out:
            diff = rec.check_out - rec.check_in
            hours = diff.total_seconds() / 3600
            duration_str = f"{hours:.2f}h"

        row = {
            "S.No": i,
            "Date": local_check_in.strftime("%d-%m-%Y"),
        }

        if not user_id:
            user_info = user_map.get(rec.user_id, {"name": "Unknown", "email": "Unknown"})
            row["Employee Name"] = user_info["name"]
            row["Email"] = user_info["email"]

        row.update({
            "Check In": local_check_in.strftime("%H:%M:%S"),
            "Check Out": local_check_out.strftime("%H:%M:%S") if local_check_out else "N/A",
            "Duration": duration_str,
            "Status": rec.status.upper(),
            "Address (In)": rec.address_in or "",
            "Address (Out)": rec.address_out or "",
            "Remarks": rec.remarks or ""
        })
        rows.append(row)

    df = pd.DataFrame(rows)
    # Convert all column names to UPPERCASE
    df.columns = [str(c).upper() for c in df.columns]
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Attendance", index=False)
    output.seek(0)
    return output


async def calculate_payroll_data(current_user: User, year: int, month: int, employee_id: Optional[str] = None):
    from app.models.system_settings import SystemSettings
    from app.models.company import Company
    from app.services.reward_service import calculate_employee_performance

    # Get settings
    settings = None
    if current_user.company_id:
        settings = await SystemSettings.find_one(SystemSettings.company_id == current_user.company_id)
    if not settings:
        settings = await SystemSettings.find_one({"singleton_id": "default"})
    if not settings:
        settings = SystemSettings()

    # Get start and end dates in UTC
    import calendar
    
    start_date = datetime(year, month, 1, 0, 0, 0, tzinfo=timezone.utc)
    _, last_day = calendar.monthrange(year, month)
    end_date = datetime(year, month, last_day, 23, 59, 59, 999999, tzinfo=timezone.utc)

    # Get employees based on hierarchy and filter
    from app.services import user_service
    from app.models.role import BaseArchetype
    arch = current_user.role_archetype or current_user.role
    if arch in [BaseArchetype.EMPLOYEE, BaseArchetype.CONTRACTOR, UserRole.EMPLOYEE]:
        employees = [current_user]
    elif employee_id:
        target_uid = PydanticObjectId(employee_id)
        if target_uid == current_user.id:
            employees = [current_user]
        else:
            subordinates = await user_service.get_all_employees(current_user)
            subordinate_ids = {emp.id for emp in subordinates}
            if target_uid not in subordinate_ids:
                employees = []
            else:
                emp_obj = await User.get(target_uid)
                employees = [emp_obj] if emp_obj else []
    else:
        employees = await user_service.get_all_employees(current_user)

    employee_ids = [emp.id for emp in employees]

    # Get all tasks for this month for these employees
    tasks = await Task.find(
        {"assigned_to": {"$in": employee_ids}},
        {"$or": [
            {"completed_at": {"$gte": start_date, "$lte": end_date}},
            {"completed_at": None, "deadline": {"$gte": start_date, "$lte": end_date}}
        ]}
     ).to_list()

    # Group tasks by assigned_to
    tasks_by_user = {}
    for task in tasks:
        uid = str(task.assigned_to)
        if uid not in tasks_by_user:
            tasks_by_user[uid] = []
        tasks_by_user[uid].append(task)

    # Get all attendance records for this month for these employees
    attendance_records = await Attendance.find(
        {"user_id": {"$in": employee_ids}},
        Attendance.check_in >= start_date,
        Attendance.check_in <= end_date
    ).to_list()

    # Group attendance by user_id
    attendance_by_user = {}
    for att in attendance_records:
        uid = str(att.user_id)
        if uid not in attendance_by_user:
            attendance_by_user[uid] = []
        attendance_by_user[uid].append(att)

    # Fetch all companies in a map for start time settings
    companies = await Company.find_all().to_list()
    company_map = {str(c.id): c for c in companies}

    # Determine scheduled weekdays
    today = datetime.utcnow() + timedelta(hours=5, minutes=30)
    cal = calendar.Calendar()
    all_weekdays = []
    for d in cal.itermonthdates(year, month):
        if d.month == month and d.year == year and d.weekday() < 5:
            all_weekdays.append(d)
            
    scheduled_days = []
    for d in all_weekdays:
        if year < today.year or (year == today.year and month < today.month):
            scheduled_days.append(d)
        elif year == today.year and month == today.month:
            if d <= today.date():
                scheduled_days.append(d)
    if not scheduled_days:
        scheduled_days = all_weekdays

    payroll_records = []
    for emp in employees:
        emp_join_date = emp.created_at.date() if hasattr(emp, 'created_at') and emp.created_at else None
        emp_scheduled_days = scheduled_days
        if emp_join_date:
            emp_scheduled_days = [d for d in scheduled_days if d >= emp_join_date]

        # Map attendance records by date
        att_by_date = {}
        for att in attendance_by_user.get(str(emp.id), []):
            local_ci = att.check_in + timedelta(hours=5, minutes=30)
            att_by_date[local_ci.date()] = att

        total_attendance_points = 0.0
        present_count = 0
        late_count = 0
        absent_count = 0

        for day in emp_scheduled_days:
            att = att_by_date.get(day)
            if att:
                company = company_map.get(str(emp.company_id)) if emp.company_id else None
                work_start_str = company.work_start_time if company else "09:00 AM"
                work_start = parse_time_string(work_start_str)

                local_ci = att.check_in + timedelta(hours=5, minutes=30)
                start_dt = local_ci.replace(hour=work_start.hour, minute=work_start.minute, second=0, microsecond=0)
                diff_minutes = (local_ci - start_dt).total_seconds() / 60.0

                if diff_minutes <= 0:
                    pts = settings.attendance_impact.get("present", 1.0)
                    present_count += 1
                elif diff_minutes <= 30:
                    pts = settings.attendance_impact.get("late_under_30", 0.75)
                    late_count += 1
                else:
                    pts = settings.attendance_impact.get("late_over_30", 0.50)
                    late_count += 1

                # Check overtime: worked >= 9.5 hours
                if att.check_out:
                    duration_hours = (att.check_out - att.check_in).total_seconds() / 3600.0
                    if duration_hours >= 9.5:
                        pts = settings.attendance_impact.get("overtime", 1.25)

                total_attendance_points += pts
            else:
                pts = settings.attendance_impact.get("unexcused", -1.0)
                total_attendance_points += pts
                absent_count += 1

        # Attendance reliability %
        if emp_scheduled_days:
            reliability_score = (total_attendance_points / len(emp_scheduled_days)) * 100.0
            reliability_score = max(0.0, min(100.0, reliability_score))
        else:
            reliability_score = 100.0

        # Attendance bonus
        base_sal = getattr(emp, "base_salary", 30000.0)
        attendance_bonus = 0.05 * base_sal if reliability_score >= (settings.attendance_bonus_threshold * 100.0) else 0.0

        # Calculate Task Points using reward service
        emp_tasks = tasks_by_user.get(str(emp.id), [])
        perf = await calculate_employee_performance(
            user_id=emp.id,
            start_date=start_date,
            end_date=end_date,
            tasks_list=emp_tasks
        )

        earned_task_points = perf["total_earned_points"]
        total_possible_points = perf["total_possible_points"]
        performance_pct = perf["performance_percent"]
        late_tasks_4_days_plus = perf["overdue_count"]
        applied_deduction = perf["deduction_applied"]
        incentive_mult = perf["incentive_multiplier"]

        # Performance incentive
        incentive_pool = 0.25 * base_sal
        performance_incentive = incentive_mult * incentive_pool

        # Gross Pay
        gross_pay = base_sal + attendance_bonus + performance_incentive

        payroll_records.append({
            "employee_id": str(emp.id),
            "name": emp.name,
            "email": emp.email,
            "role": emp.role.value if hasattr(emp.role, 'value') else str(emp.role),
            "base_salary": base_sal,
            "scheduled_days": len(emp_scheduled_days),
            "present_days": present_count,
            "late_days": late_count,
            "absent_days": absent_count,
            "attendance_points": round(total_attendance_points, 2),
            "attendance_reliability": round(reliability_score, 2),
            "attendance_bonus": round(attendance_bonus, 2),
            "earned_task_points": round(earned_task_points, 2),
            "total_possible_points": round(total_possible_points, 2),
            "performance_percentage": round(performance_pct, 2),
            "late_tasks_backlog": late_tasks_4_days_plus,
            "applied_backlog_deduction": round(applied_deduction, 2),
            "performance_incentive": round(performance_incentive, 2),
            "gross_pay": round(gross_pay, 2)
        })

    return payroll_records


async def generate_payroll_excel(current_user: User, year: int, month: int, employee_id: Optional[str] = None) -> BytesIO:
    records = await calculate_payroll_data(current_user, year, month, employee_id)
    
    # Create DataFrame
    df = pd.DataFrame(records)
    
    # Check if empty
    if df.empty:
        df_filtered = pd.DataFrame(columns=[
            "EMPLOYEE NAME", "EMAIL", "ROLE", "BASE SALARY", "SCHEDULED DAYS", 
            "PRESENT DAYS", "LATE DAYS", "ABSENT DAYS", "ATTENDANCE RELIABILITY %", 
            "ATTENDANCE BONUS", "EARNED POINTS", "POSSIBLE POINTS", "PERFORMANCE %", 
            "LATE BACKLOGS (4D+)", "PERFORMANCE INCENTIVE", "GROSS PAY"
        ])
    else:
        # Re-order and rename columns to be user-friendly and uppercase
        columns_mapping = {
            "name": "EMPLOYEE NAME",
            "email": "EMAIL",
            "role": "ROLE",
            "base_salary": "BASE SALARY",
            "scheduled_days": "SCHEDULED DAYS",
            "present_days": "PRESENT DAYS",
            "late_days": "LATE DAYS",
            "absent_days": "ABSENT DAYS",
            "attendance_reliability": "ATTENDANCE RELIABILITY %",
            "attendance_bonus": "ATTENDANCE BONUS",
            "earned_task_points": "EARNED POINTS",
            "total_possible_points": "POSSIBLE POINTS",
            "performance_percentage": "PERFORMANCE %",
            "late_tasks_backlog": "LATE BACKLOGS (4D+)",
            "performance_incentive": "PERFORMANCE INCENTIVE",
            "gross_pay": "GROSS PAY"
        }
        df_filtered = df[list(columns_mapping.keys())].rename(columns=columns_mapping)
    
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df_filtered.to_excel(writer, sheet_name="Payroll Report", index=False)
        
        # Get sheet to apply styles
        workbook = writer.book
        worksheet = writer.sheets["Payroll Report"]
        
        # Enable grid lines explicitly
        worksheet.views.sheetView[0].showGridLines = True
        
        # Import styling tools
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Palette: Dark Slate Blue & Ice Blue accents
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid") # Dark Slate
        
        data_font = Font(name="Segoe UI", size=10)
        bold_data_font = Font(name="Segoe UI", size=10, bold=True)
        accent_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Slate 100
        
        border_side = Side(border_style="thin", color="CBD5E1") # Light gray borders
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
        
        # Apply header styling
        for col_idx in range(1, len(df_filtered.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            
        # Row heights
        worksheet.row_dimensions[1].height = 28
        
        # Style data cells
        for row_idx in range(2, len(df_filtered) + 2):
            worksheet.row_dimensions[row_idx].height = 20
            # alternating row colors or highlight gross pay
            for col_idx in range(1, len(df_filtered.columns) + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = border
                
                # Check column type to set alignment and number format
                col_name = df_filtered.columns[col_idx - 1]
                
                if col_name in ["EMPLOYEE NAME", "EMAIL", "ROLE"]:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif col_name in ["SCHEDULED DAYS", "PRESENT DAYS", "LATE DAYS", "ABSENT DAYS", "LATE BACKLOGS (4D+)"]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.number_format = "#,##0"
                elif col_name in ["ATTENDANCE RELIABILITY %", "PERFORMANCE %"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "0.00\"%\""
                elif col_name in ["BASE SALARY", "ATTENDANCE BONUS", "PERFORMANCE INCENTIVE", "GROSS PAY"]:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "$#,##0.00"
                    if col_name == "GROSS PAY":
                        cell.font = bold_data_font
                        cell.fill = accent_fill
                        
        # Auto-adjust column widths
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                # handle formatting strings visual length
                if cell.number_format and "$" in cell.number_format:
                    val_str = "$" + val_str
                if len(val_str) > max_len:
                    max_len = len(val_str)
            worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
    output.seek(0)
    return output
